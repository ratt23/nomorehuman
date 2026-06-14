import os
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def process_harmonization(main_path: str, dict_path: str, config: dict, output_dir: str) -> tuple[str, dict]:
    """
    Processor function for aligning and reconciling Excel sheets.
    Supports matching based on: code, name, or both (dual key).
    Supports highlighting changes and excluding price/numeric columns from formatting alterations.
    """
    main_sheet = config["mainSheet"]
    dict_sheet = config["dictSheet"]
    match_basis = config["matchBasis"]
    keys = config["keys"]
    
    # Load sheets into Pandas
    main_df = pd.read_excel(main_path, sheet_name=main_sheet)
    dict_df = pd.read_excel(dict_path, sheet_name=dict_sheet)
    
    # Parse key columns
    if match_basis == 'both':
        main_key1 = keys["mainKey1"]
        dict_key1 = keys["dictKey1"]
        main_key2 = keys["mainKey2"]
        dict_key2 = keys["dictKey2"]
    else:
        main_key = keys["mainKey"]
        dict_key = keys["dictKey"]
        
    rows_processed = len(main_df)
    rows_updated = 0
    rows_deleted = 0
    
    modified_cells = [] # list of (df_idx, col_idx)
    rows_to_drop = []
    
    # Prepare price columns list (case-insensitive, trimmed)
    price_cols_lower = [c.lower().strip() for c in config.get("priceColumns", [])]
    
    for idx, row in main_df.iterrows():
        # Find match condition
        if match_basis == 'both':
            val_kode = row.get(main_key1)
            val_nama = row.get(main_key2)
            
            # Clean values for comparison
            dict_col1 = dict_df[dict_key1].astype(str).str.strip().str.lower()
            dict_col2 = dict_df[dict_key2].astype(str).str.strip().str.lower()
            
            match_cond = (dict_col1 == str(val_kode).strip().lower()) & (dict_col2 == str(val_nama).strip().lower())
        else:
            val = row.get(main_key)
            dict_col = dict_df[dict_key].astype(str).str.strip().str.lower()
            match_cond = (dict_col == str(val).strip().lower())
            
        matches = dict_df[match_cond]
        
        if not matches.empty:
            match_row = matches.iloc[0]
            
            if config.get("deleteMatched", False):
                rows_to_drop.append(idx)
                rows_deleted += 1
            else:
                updated_any = False
                for mapping in config.get("mappings", []):
                    main_col = mapping["mainCol"]
                    dict_col_name = mapping["dictCol"]
                    
                    # Protect pricing columns from automated text/casing manipulation
                    if main_col.lower().strip() in price_cols_lower:
                        continue
                        
                    new_val = match_row.get(dict_col_name)
                    
                    if pd.notna(new_val):
                        old_val = main_df.at[idx, main_col]
                        
                        # Compare string representations to detect change
                        if str(old_val).strip().lower() != str(new_val).strip().lower():
                            # Assign new value directly (preserves float/int types)
                            main_df.at[idx, main_col] = new_val
                            updated_any = True
                            
                            # Record the modified cell using index and 1-based column offset
                            col_idx = list(main_df.columns).index(main_col) + 1
                            modified_cells.append((idx, col_idx))
                
                if updated_any:
                    rows_updated += 1
                    
    # Drop deleted rows
    if rows_to_drop:
        main_df.drop(rows_to_drop, inplace=True)
        
    # Sort dataframe if requested
    sort_col = config.get("sortColumn")
    if sort_col and sort_col in main_df.columns:
        ascending = config.get("sortOrder", "asc") == "asc"
        main_df.sort_values(by=sort_col, ascending=ascending, inplace=True)
        
    # Generate target output path
    output_filename = f"harmonized_{int(time.time())}.xlsx"
    result_path = os.path.join(output_dir, output_filename)
    
    # Save the dataframe back to Excel (without index column)
    main_df.to_excel(result_path, sheet_name=main_sheet, index=False)
    
    # Determine new Excel row mapping for remaining indices (headers in row 1, data starts in row 2)
    idx_to_excel_row = {idx: (output_idx + 2) for output_idx, idx in enumerate(main_df.index)}
    
    # Apply highlights to changed cells using openpyxl
    if config.get("highlightChanges", False) and modified_cells:
        wb = load_workbook(result_path)
        ws = wb[main_sheet]
        
        yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        for idx, col_idx in modified_cells:
            # Check if row was not deleted
            if idx in idx_to_excel_row:
                row_num = idx_to_excel_row[idx]
                cell = ws.cell(row=row_num, column=col_idx)
                cell.fill = yellow_fill
                
        wb.save(result_path)
        wb.close()
        
    diagnostics = {
        "rowsProcessed": rows_processed,
        "rowsUpdated": rows_updated,
        "rowsDeleted": rows_deleted
    }
    
    return result_path, diagnostics
