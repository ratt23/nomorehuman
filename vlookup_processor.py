import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font

def perform_vlookup(main_file_path, template_file_path, mappings):
    main_wb = load_workbook(main_file_path)
    main_sheet = main_wb.worksheets[0]
    
    template_wb = load_workbook(template_file_path, read_only=True)
    template_sheet = template_wb.worksheets[0]
    
    template_key_col = -1
    template_value_col = -1
    
    template_header_row = [cell.value for cell in next(template_sheet.iter_rows(max_row=1))]
    for col_idx, header in enumerate(template_header_row, 1):
        if header and str(header).strip() == mappings.get('templateKey'):
            template_key_col = col_idx
        if header and str(header).strip() == mappings.get('valueToGet'):
            template_value_col = col_idx
            
    if template_key_col == -1 or template_value_col == -1:
        raise Exception('Kolom kunci atau kolom nilai tidak ditemukan di file template.')
        
    lookup_map = {}
    for row in template_sheet.iter_rows(min_row=2):
        key_val = row[template_key_col - 1].value
        key = str(key_val).strip() if key_val is not None else ''
        value = row[template_value_col - 1].value
        if key:
            lookup_map[key] = value
            
    template_wb.close()
            
    main_key_col = -1
    main_header_row = [cell.value for cell in next(main_sheet.iter_rows(max_row=1))]
    for col_idx, header in enumerate(main_header_row, 1):
        if header and str(header).strip() == mappings.get('mainKey'):
            main_key_col = col_idx
            
    if main_key_col == -1:
        raise Exception('Kolom kunci tidak ditemukan di file utama.')
        
    diagnostics = {
        'summary': {
            'totalRowsInMainFile': main_sheet.max_row - 1,
            'totalRowsMatched': 0,
            'totalRowsUnmatched': 0,
        },
        'unmatchedRowsSample': [],
        'matchedRowsSample': [],
        'SAMPLE_SIZE': 10
    }
    
    new_column_index = main_sheet.max_column + 1
    new_header_cell = main_sheet.cell(row=1, column=new_column_index, value=mappings.get('valueToGet'))
    new_header_cell.font = Font(bold=True)
    
    for row_num in range(2, main_sheet.max_row + 1):
        key_cell_val = main_sheet.cell(row=row_num, column=main_key_col).value
        key_to_find = str(key_cell_val).strip() if key_cell_val is not None else ''
        if key_to_find in lookup_map:
            diagnostics['summary']['totalRowsMatched'] += 1
            found_value = lookup_map[key_to_find]
            main_sheet.cell(row=row_num, column=new_column_index, value=found_value)
            
            if len(diagnostics['matchedRowsSample']) < diagnostics['SAMPLE_SIZE']:
                diagnostics['matchedRowsSample'].append({
                    'key': key_to_find,
                    'value': found_value
                })
        else:
            diagnostics['summary']['totalRowsUnmatched'] += 1
            if len(diagnostics['unmatchedRowsSample']) < diagnostics['SAMPLE_SIZE']:
                diagnostics['unmatchedRowsSample'].append({'key': key_to_find})
                
    total_in_main = diagnostics['summary']['totalRowsInMainFile']
    diagnostics['summary']['matchPercentage'] = (diagnostics['summary']['totalRowsMatched'] / total_in_main * 100) if total_in_main > 0 else 0
    
    output_dir = os.path.join(os.path.dirname(__file__), 'output')
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime('%Y-%m-%dT%H-%M-%S')
    out_file_name = f"VLOOKUP_Result_{timestamp}.xlsx"
    out_path = os.path.join(output_dir, out_file_name)
    
    main_wb.save(out_path)
    main_wb.close()
    return out_path, diagnostics
