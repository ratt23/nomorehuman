import os
import shutil
import json
import re
import time
import platform
import psutil
from datetime import datetime
from fastapi import FastAPI, File, UploadFile, Form, HTTPException, Request, Response
from fastapi.staticfiles import StaticFiles
from fastapi.responses import JSONResponse, HTMLResponse, RedirectResponse
from openpyxl import load_workbook

# Import modules
from process_excel import process_and_diagnose_sheet, create_final_excel, parse_price
from generate_pdf import generate_pdf_from_data
import report_constructor_engine
import vlookup_processor
import validate_excel
from fastapi.middleware.cors import CORSMiddleware

# Track server start time for uptime calculation
SERVER_START_TIME = time.time()

app = FastAPI(title="no more human!! Backend")

# Enable CORS for hybrid deploy (allows Netlify frontend to communicate with Render backend)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

import base64
import tempfile

IS_SERVERLESS = "NETLIFY" in os.environ or "LAMBDA_TASK_ROOT" in os.environ

if IS_SERVERLESS:
    UPLOAD_DIR = os.path.join(tempfile.gettempdir(), "uploads")
    OUTPUT_DIR = os.path.join(tempfile.gettempdir(), "output")
else:
    UPLOAD_DIR = "uploads"
    OUTPUT_DIR = "output"

for d in [UPLOAD_DIR, OUTPUT_DIR]:
    os.makedirs(d, exist_ok=True)

if not IS_SERVERLESS:
    os.makedirs("public", exist_ok=True)

MAX_FILE_SIZE = 10 * 1024 * 1024 # 10 MB

def validate_excel_mime(file: UploadFile):
    allowed_extensions = ['.xlsx', '.xls']
    _, ext = os.path.splitext(file.filename)
    if ext.lower() not in allowed_extensions:
        raise HTTPException(
            status_code=400, 
            detail=f"Format file tidak valid. Aplikasi hanya menerima file Excel (.xlsx, .xls), bukan {file.filename}."
        )
    
    try:
        file.file.seek(0, os.SEEK_END)
        size = file.file.tell()
        file.file.seek(0)
        if size > MAX_FILE_SIZE:
            raise HTTPException(
                status_code=400,
                detail=f"Ukuran berkas melebihi batas 10 MB (Ukuran berkas Anda: {size / (1024*1024):.1f} MB)."
            )
    except Exception as e:
        if isinstance(e, HTTPException):
            raise e
        pass

# Helper function for temporary file handling
def save_temp_file(file: UploadFile) -> str:
    temp_path = os.path.join(UPLOAD_DIR, f"{int(datetime.now().timestamp())}_{file.filename}")
    with open(temp_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    return temp_path

def delete_file(path: str):
    if os.path.exists(path):
        try:
            os.remove(path)
        except Exception:
            pass

@app.post("/inspect-file")
async def inspect_file(file: UploadFile = File(...)):
    validate_excel_mime(file)
    temp_path = save_temp_file(file)
    try:
        wb = load_workbook(temp_path, read_only=True)
        inspection_data = {}
        MAX_ROWS_TO_INSPECT = 5000
        
        for sheet in wb.worksheets:
            # Check row count
            if sheet.max_row is not None and sheet.max_row <= 1:
                continue
            
            # Load normal workbook for each sheet (to read values safely)
            # Actually, read_only workbook works too
            headers_with_col = []
            unique_values_per_column = {}
            
            # Get headers from first row
            header_row = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
            for col_num, val in enumerate(header_row, 1):
                if val is not None:
                    header_text = str(val).strip()
                    if header_text:
                        headers_with_col.append({'text': header_text, 'col': col_num})
                        unique_values_per_column[header_text] = set()
            
            rows_inspected = 0
            for row in sheet.iter_rows(min_row=2):
                if rows_inspected >= MAX_ROWS_TO_INSPECT:
                    break
                
                # Collect values
                for h_info in headers_with_col:
                    col_idx = h_info['col']
                    cell_val = row[col_idx - 1].value
                    if cell_val is not None:
                        val_str = str(cell_val).strip()
                        if val_str:
                            unique_values_per_column[h_info['text']].add(val_str)
                rows_inspected += 1
                
            # Convert sets to sorted lists
            formatted_unique = {}
            for header, val_set in unique_values_per_column.items():
                formatted_unique[header] = sorted(list(val_set))
                
            inspection_data[sheet.title] = {
                'headers': [h['text'] for h in headers_with_col],
                'uniqueValuesPerColumn': formatted_unique
            }
        
        wb.close()
        return {'ok': True, 'sheets': inspection_data}
    except Exception as e:
        return JSONResponse(status_code=400, content={'ok': False, 'error': str(e)})
    finally:
        delete_file(temp_path)

@app.post("/process-file")
async def process_file(
    file: UploadFile = File(...),
    mappings: str = Form(...),
    classMap: str = Form(...),
    sheet: str = Form(...),
    filterConfig: str = Form(...)
):
    validate_excel_mime(file)
    temp_path = save_temp_file(file)
    try:
        mappings_data = json.loads(mappings)
        class_map_data = json.loads(classMap)
        filter_config_data = json.loads(filterConfig)
        
        res = process_and_diagnose_sheet(
            temp_path, mappings_data, class_map_data, sheet, filter_config_data
        )
        
        all_accepted_rows = res['allAcceptedRows']
        if len(all_accepted_rows) == 0:
            raise Exception("Tidak ada data yang diproses setelah filter diterapkan.")
            
        final_xlsx_path = create_final_excel(all_accepted_rows, OUTPUT_DIR)
        pdf_path = final_xlsx_path.replace('.xlsx', '.pdf')
        
        generate_pdf_from_data(all_accepted_rows, pdf_path)
        
        # Read and encode output files to Base64
        with open(final_xlsx_path, "rb") as f:
            excel_base64 = base64.b64encode(f.read()).decode('utf-8')
        with open(pdf_path, "rb") as f:
            pdf_base64 = base64.b64encode(f.read()).decode('utf-8')
            
        if IS_SERVERLESS:
            delete_file(final_xlsx_path)
            delete_file(pdf_path)
            
        return {
            'ok': True,
            'message': 'File berhasil diproses!',
            'excel': f"/output/{os.path.basename(final_xlsx_path)}",
            'pdf': f"/output/{os.path.basename(pdf_path)}",
            'excelData': excel_base64,
            'pdfData': pdf_base64,
            'excelFilename': os.path.basename(final_xlsx_path),
            'pdfFilename': os.path.basename(pdf_path),
            'diagnostics': {
                'summary': res['summary'],
                'rejectedRowsSample': res['rejectedRowsSample'],
                'acceptedRowsSample': res['acceptedRowsSample']
            }
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={'ok': False, 'error': str(e)})
    finally:
        delete_file(temp_path)

def compare_files(original_path, processed_path, mappings, class_map, selected_sheet):
    original_wb = load_workbook(original_path, data_only=True)
    processed_wb = load_workbook(processed_path, data_only=True)
    
    if selected_sheet not in original_wb.sheetnames:
        raise Exception(f'Sheet "{selected_sheet}" tidak ditemukan di file original')
        
    original_sheet = original_wb[selected_sheet]
    processed_sheet = processed_wb.worksheets[0]
    
    results = {
        'summary': { 'itemsCompared': 0, 'priceMatches': 0, 'priceMismatches': 0 },
        'priceComparison': []
    }
    
    original_headers = {}
    original_header_row = [cell.value for cell in next(original_sheet.iter_rows(max_row=1))]
    for col_idx, val in enumerate(original_header_row, 1):
        if val is not None:
            original_headers[str(val).strip()] = col_idx
            
    col_code = original_headers.get(mappings.get('kode'))
    col_name = original_headers.get(mappings.get('nama'))
    col_class = original_headers.get(mappings.get('kelas'))
    col_price = original_headers.get(mappings.get('harga'))
    
    if not all([col_code, col_name, col_class, col_price]):
        raise Exception('Mapping kolom tidak valid untuk file original')
        
    processed_headers = {}
    processed_header_row = [cell.value for cell in next(processed_sheet.iter_rows(max_row=1))]
    for col_idx, val in enumerate(processed_header_row, 1):
        if val is not None:
            processed_headers[str(val).strip()] = col_idx
            
    class_columns = ['OPD', 'ED', 'KELAS 3', 'KELAS 2', 'KELAS 1', 'VIP', 'VVIP']
    class_to_column_map = {}
    for class_name in class_columns:
        class_to_column_map[class_name] = processed_headers.get(class_name)
        
    original_data_map = {}
    
    for row_num in range(2, original_sheet.max_row + 1):
        code_val = original_sheet.cell(row=row_num, column=col_code).value
        name_val = original_sheet.cell(row=row_num, column=col_name).value
        code = str(code_val).strip() if code_val is not None else ''
        name = str(name_val).strip() if name_val is not None else ''
        if not code or not name:
            continue
            
        kelas_val = original_sheet.cell(row=row_num, column=col_class).value
        kelas = str(kelas_val).strip().upper() if kelas_val is not None else ''
        price = parse_price(original_sheet.cell(row=row_num, column=col_price).value)
        
        key = f"{code}|{name}".upper()
        if key not in original_data_map:
            original_data_map[key] = { 'code': code, 'name': name, 'prices': {} }
            
        target_class = class_map.get(kelas, kelas)
        if target_class and target_class != 'ignore':
            original_data_map[key]['prices'][target_class] = price
            
    code_proc_col = processed_headers.get('Kode')
    name_proc_col = processed_headers.get('Nama Pemeriksaan')
    
    if not code_proc_col or not name_proc_col:
        raise Exception('Kolom "Kode" atau "Nama Pemeriksaan" tidak ditemukan di file processed')
        
    for row_num in range(2, processed_sheet.max_row + 1):
        proc_code_val = processed_sheet.cell(row=row_num, column=code_proc_col).value
        proc_name_val = processed_sheet.cell(row=row_num, column=name_proc_col).value
        processed_code = str(proc_code_val).strip() if proc_code_val is not None else ''
        processed_name = str(proc_name_val).strip() if proc_name_val is not None else ''
        
        if not processed_code or not processed_name:
            continue
            
        key = f"{processed_code}|{processed_name}".upper()
        original_item = original_data_map.get(key)
        
        if not original_item:
            results['priceComparison'].append({
                'code': processed_code,
                'name': processed_name,
                'status': 'NOT_FOUND',
                'message': 'Item tidak ditemukan di file original'
            })
            continue
            
        results['summary']['itemsCompared'] += 1
        comparison = {
            'code': processed_code,
            'name': processed_name,
            'status': 'MATCH',
            'details': []
        }
        
        all_match = True
        for class_name in class_columns:
            proc_col_idx = class_to_column_map.get(class_name)
            if not proc_col_idx:
                continue
                
            processed_price = parse_price(processed_sheet.cell(row=row_num, column=proc_col_idx).value)
            original_price = original_item['prices'].get(class_name)
            
            is_match = processed_price == original_price
            comparison['details'].append({
                'class': class_name,
                'originalPrice': original_price,
                'processedPrice': processed_price,
                'match': is_match
            })
            
            if not is_match:
                all_match = False
                
        if all_match:
            comparison['status'] = 'MATCH'
            results['summary']['priceMatches'] += 1
        else:
            comparison['status'] = 'MISMATCH'
            results['summary']['priceMismatches'] += 1
            
        results['priceComparison'].append(comparison)
        
    items_compared = results['summary']['itemsCompared']
    results['summary']['matchPercentage'] = (results['summary']['priceMatches'] / items_compared * 100) if items_compared > 0 else 0
    
    original_wb.close()
    processed_wb.close()
    return results

@app.post("/double-check-tarif")
async def double_check_tarif(
    originalFile: UploadFile = File(...),
    processedFile: UploadFile = File(...),
    mappings: str = Form(...),
    classMap: str = Form(...),
    selectedSheet: str = Form(...)
):
    orig_path = save_temp_file(originalFile)
    proc_path = save_temp_file(processedFile)
    try:
        mappings_data = json.loads(mappings)
        class_map_data = json.loads(classMap)
        
        comparison = compare_files(
            orig_path, proc_path, mappings_data, class_map_data, selectedSheet
        )
        return {'ok': True, 'message': 'Double check harga selesai!', 'comparison': comparison}
    except Exception as e:
        return JSONResponse(status_code=500, content={'ok': False, 'error': str(e)})
    finally:
        delete_file(orig_path)
        delete_file(proc_path)

@app.post("/inspect-source-file")
async def inspect_source_file_endpoint(
    file: UploadFile = File(...),
    config: str = Form(None)
):
    validate_excel_mime(file)
    temp_path = save_temp_file(file)
    try:
        config_data = json.loads(config) if config else None
        res = report_constructor_engine.inspect_source_file(temp_path, config_data)
        return {'ok': True, **res}
    except Exception as e:
        return JSONResponse(status_code=400, content={'ok': False, 'error': str(e)})
    finally:
        delete_file(temp_path)

@app.post("/inspect-template")
async def inspect_template_endpoint(templateFile: UploadFile = File(...)):
    validate_excel_mime(templateFile)
    temp_path = save_temp_file(templateFile)
    try:
        headers = report_constructor_engine.inspect_template_file(temp_path)
        return {'ok': True, 'headers': headers}
    except Exception as e:
        return JSONResponse(status_code=400, content={'ok': False, 'error': str(e)})
    finally:
        delete_file(temp_path)

@app.post("/build-report")
async def build_report_endpoint(
    file: UploadFile = File(...),
    config: str = Form(...)
):
    validate_excel_mime(file)
    temp_path = save_temp_file(file)
    try:
        config_data = json.loads(config)
        result_path, diagnostics = report_constructor_engine.build_report(temp_path, config_data)
        
        # Read and encode to Base64
        with open(result_path, "rb") as f:
            excel_base64 = base64.b64encode(f.read()).decode('utf-8')
            
        if IS_SERVERLESS:
            delete_file(result_path)
            
        return {
            'ok': True,
            'message': 'Laporan berhasil dibuat!',
            'excel': f"/output/{os.path.basename(result_path)}",
            'excelData': excel_base64,
            'excelFilename': os.path.basename(result_path),
            'diagnostics': diagnostics
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={'ok': False, 'error': str(e)})
    finally:
        delete_file(temp_path)

# --- VLOOKUP ENDPOINTS (NOW INTEGRATED!) ---

@app.post("/vlookup-inspect")
async def vlookup_inspect_endpoint(
    mainFile: UploadFile = File(...),
    templateFile: UploadFile = File(...)
):
    validate_excel_mime(mainFile)
    validate_excel_mime(templateFile)
    main_path = save_temp_file(mainFile)
    temp_path = save_temp_file(templateFile)
    try:
        # Load workbooks to read headers
        m_wb = load_workbook(main_path, read_only=True)
        t_wb = load_workbook(temp_path, read_only=True)
        
        main_headers = [str(cell.value).strip() for cell in next(m_wb.worksheets[0].iter_rows(max_row=1)) if cell.value is not None]
        template_headers = [str(cell.value).strip() for cell in next(t_wb.worksheets[0].iter_rows(max_row=1)) if cell.value is not None]
        
        m_wb.close()
        t_wb.close()
        
        return {
            'ok': True,
            'mainHeaders': main_headers,
            'templateHeaders': template_headers
        }
    except Exception as e:
        return JSONResponse(status_code=400, content={'ok': False, 'error': str(e)})
    finally:
        delete_file(main_path)
        delete_file(temp_path)

@app.post("/vlookup-process")
async def vlookup_process_endpoint(
    mainFile: UploadFile = File(...),
    templateFile: UploadFile = File(...),
    mappings: str = Form(...)
):
    validate_excel_mime(mainFile)
    validate_excel_mime(templateFile)
    main_path = save_temp_file(mainFile)
    temp_path = save_temp_file(templateFile)
    try:
        mappings_data = json.loads(mappings)
        result_path, diagnostics = vlookup_processor.perform_vlookup(
            main_path, temp_path, mappings_data
        )
        
        # Read and encode to Base64
        with open(result_path, "rb") as f:
            excel_base64 = base64.b64encode(f.read()).decode('utf-8')
            
        if IS_SERVERLESS:
            delete_file(result_path)
            
        return {
            'ok': True,
            'excel': f"/output/{os.path.basename(result_path)}",
            'excelData': excel_base64,
            'excelFilename': os.path.basename(result_path),
            'diagnostics': diagnostics
        }
    except Exception as e:
        return JSONResponse(status_code=500, content={'ok': False, 'error': str(e)})
    finally:
        delete_file(main_path)
        delete_file(temp_path)

ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "nomorehuman2026")

def get_login_page_html(error: str = None):
    error_html = f'<div class="error-msg">{error}</div>' if error else ''
    return f"""
    <!doctype html>
    <html lang="id">
    <head>
        <meta charset="utf-8"/>
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Login - NO MORE HUMAN</title>
        <link rel="preconnect" href="https://fonts.googleapis.com">
        <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
        <link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700&family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
        <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.1/css/all.min.css">
        <style>
            :root {{
                --bg-color: #080B12;
                --panel-bg: #101827;
                --primary-color: #0066FF;
                --text-color: #F3F4F6;
                --text-secondary: #9CA3AF;
                --border-color: rgba(255, 255, 255, 0.08);
                --danger-color: #EF4444;
            }}
            body {{
                font-family: 'Plus Jakarta Sans', sans-serif;
                background-color: var(--bg-color);
                color: var(--text-color);
                margin: 0;
                display: flex;
                align-items: center;
                justify-content: center;
                height: 100vh;
                overflow: hidden;
            }}
            .login-container {{
                background: var(--panel-bg);
                border: 1px solid var(--border-color);
                border-radius: 12px;
                padding: 40px;
                width: 100%;
                max-width: 400px;
                box-shadow: 0 20px 40px rgba(0, 0, 0, 0.5), 0 0 20px rgba(0, 102, 255, 0.05);
                text-align: center;
                box-sizing: border-box;
            }}
            .logo-wrap {{
                font-size: 2.5rem;
                color: var(--primary-color);
                margin-bottom: 15px;
            }}
            h2 {{
                margin: 0 0 5px 0;
                font-size: 1.8rem;
                font-weight: 700;
                letter-spacing: -0.025em;
            }}
            .tagline {{
                color: var(--primary-color);
                font-size: 0.95rem;
                font-weight: 600;
                font-family: monospace;
                margin: 0 0 20px 0;
            }}
            .sub-desc {{
                color: var(--text-secondary);
                font-size: 0.85rem;
                margin: 0 0 25px 0;
                opacity: 0.8;
            }}
            .form-group {{
                margin-bottom: 20px;
                text-align: left;
            }}
            label {{
                display: block;
                font-size: 0.85rem;
                font-weight: 600;
                margin-bottom: 8px;
                color: var(--text-secondary);
            }}
            input[type="password"] {{
                width: 100%;
                padding: 12px;
                background: rgba(255, 255, 255, 0.05);
                border: 1px solid var(--border-color);
                border-radius: 6px;
                color: var(--text-color);
                font-size: 1rem;
                box-sizing: border-box;
                transition: border-color 0.2s, background-color 0.2s;
            }}
            input[type="password"]:focus {{
                border-color: var(--primary-color);
                outline: none;
                background: rgba(0, 102, 255, 0.05);
            }}
            .btn {{
                width: 100%;
                padding: 12px;
                background: var(--primary-color);
                border: none;
                border-radius: 6px;
                color: white;
                font-size: 1rem;
                font-weight: 600;
                cursor: pointer;
                transition: background-color 0.2s;
            }}
            .btn:hover {{
                background-color: #0052cc;
            }}
            .error-msg {{
                color: var(--danger-color);
                font-size: 0.85rem;
                font-weight: 600;
                margin-bottom: 20px;
                padding: 10px;
                background: rgba(239, 68, 68, 0.1);
                border: 1px solid rgba(239, 68, 68, 0.2);
                border-radius: 6px;
            }}
        </style>
    </head>
    <body>
        <div class="login-container">
            <div class="logo-wrap"><i class="fas fa-microchip"></i></div>
            <h2>NO MORE HUMAN</h2>
            <p class="tagline">Humans are optional.</p>
            <p class="sub-desc">Masukkan kata sandi untuk mengakses Status Server</p>
            {error_html}
            <form action="/admin-login" method="post">
                <div class="form-group">
                    <label for="password">Kata Sandi Admin</label>
                    <input type="password" id="password" name="password" placeholder="••••••••" required autofocus>
                </div>
                <button type="submit" class="btn">Masuk <i class="fas fa-arrow-right-to-bracket"></i></button>
            </form>
        </div>
    </body>
    </html>
    """

def get_admin_dashboard_html():
    uptime_seconds = int(time.time() - SERVER_START_TIME)
    days = uptime_seconds // 86400
    hours = (uptime_seconds % 86400) // 3600
    minutes = (uptime_seconds % 3600) // 60
    seconds = uptime_seconds % 60
    uptime_str = f"{days}d {hours}h {minutes}m {seconds}s"
    
    cpu_usage = psutil.cpu_percent()
    memory = psutil.virtual_memory()
    mem_used_mb = int(memory.used / (1024 * 1024))
    mem_total_mb = int(memory.total / (1024 * 1024))
    mem_percent = memory.percent
    
    # Process memory
    process_mem_mb = int(psutil.Process(os.getpid()).memory_info().rss / (1024 * 1024))
    
    return f"""
    <!doctype html>
    <html lang="en">
    <head>
        <meta charset="utf-8"/>
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Status Server - no more human!!</title>
        <link rel="preconnect" href="https://fonts.googleapis.com">
        <link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
        <link href="https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700&family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">
        <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.5.1/css/all.min.css">
        <style>
            :root {{
                --bg-color: #080B12;
                --panel-bg: #101827;
                --primary-color: #0066FF;
                --success-color: #00D084;
                --text-color: #F3F4F6;
                --text-secondary: #9CA3AF;
                --border-color: rgba(255, 255, 255, 0.08);
            }}
            body {{
                font-family: 'Plus Jakarta Sans', sans-serif;
                background-color: var(--bg-color);
                color: var(--text-color);
                margin: 0;
                padding: 40px 20px;
                display: flex;
                flex-direction: column;
                align-items: center;
                min-height: 100vh;
                box-sizing: border-box;
            }}
            .dashboard-container {{
                width: 100%;
                max-width: 800px;
            }}
            header {{
                display: flex;
                justify-content: space-between;
                align-items: center;
                margin-bottom: 30px;
                border-bottom: 1px solid var(--border-color);
                padding-bottom: 20px;
            }}
            h2 {{
                margin: 0;
                font-size: 1.5rem;
                display: flex;
                align-items: center;
                gap: 10px;
            }}
            .btn-logout {{
                background: rgba(255, 255, 255, 0.05);
                border: 1px solid var(--border-color);
                padding: 8px 16px;
                border-radius: 6px;
                color: var(--text-color);
                text-decoration: none;
                font-size: 0.85rem;
                font-weight: 600;
                transition: background-color 0.2s;
            }}
            .btn-logout:hover {{
                background: rgba(239, 68, 68, 0.1);
                color: #EF4444;
                border-color: rgba(239, 68, 68, 0.2);
            }}
            .status-badge {{
                display: inline-flex;
                align-items: center;
                gap: 8px;
                background: rgba(0, 208, 132, 0.1);
                border: 1px solid rgba(0, 208, 132, 0.2);
                color: var(--success-color);
                padding: 6px 12px;
                border-radius: 50px;
                font-size: 0.8rem;
                font-weight: 600;
            }}
            .badge-dot {{
                width: 8px;
                height: 8px;
                background: var(--success-color);
                border-radius: 50%;
                box-shadow: 0 0 10px var(--success-color);
            }}
            .metrics-grid {{
                display: grid;
                grid-template-columns: 1fr 1fr;
                gap: 20px;
                margin-bottom: 30px;
            }}
            .metric-card {{
                background: var(--panel-bg);
                border: 1px solid var(--border-color);
                border-radius: 12px;
                padding: 25px;
                box-shadow: 0 4px 20px rgba(0,0,0,0.2);
            }}
            .metric-title {{
                color: var(--text-secondary);
                font-size: 0.85rem;
                font-weight: 600;
                margin-bottom: 15px;
                display: flex;
                align-items: center;
                gap: 8px;
            }}
            .metric-value {{
                font-size: 2rem;
                font-weight: 700;
                color: var(--text-color);
                margin-bottom: 10px;
            }}
            .progress-bar-container {{
                background: rgba(255, 255, 255, 0.05);
                height: 8px;
                border-radius: 10px;
                overflow: hidden;
            }}
            .progress-bar {{
                background: var(--primary-color);
                height: 100%;
                border-radius: 10px;
                transition: width 0.5s ease-out;
            }}
            .progress-bar.success {{
                background: var(--success-color);
            }}
            .system-details-card {{
                background: var(--panel-bg);
                border: 1px solid var(--border-color);
                border-radius: 12px;
                padding: 25px;
                box-shadow: 0 4px 20px rgba(0,0,0,0.2);
            }}
            .detail-row {{
                display: flex;
                justify-content: space-between;
                padding: 12px 0;
                border-bottom: 1px solid var(--border-color);
                font-size: 0.9rem;
            }}
            .detail-row:last-child {{
                border-bottom: none;
                padding-bottom: 0;
            }}
            .detail-row:first-child {{
                padding-top: 0;
            }}
            .detail-label {{
                color: var(--text-secondary);
                font-weight: 500;
            }}
            .detail-value {{
                font-weight: 600;
                font-family: monospace;
            }}
        </style>
        <script>
            setTimeout(() => {{
                window.location.reload();
            }}, 10000);
        </script>
    </head>
    <body>
        <div class="dashboard-container">
            <header>
                <h2><i class="fas fa-server" style="color: var(--primary-color);"></i> Panel Admin Server</h2>
                <div style="display: flex; align-items: center; gap: 15px;">
                    <div class="status-badge">
                        <span class="badge-dot"></span> <span>Online</span>
                    </div>
                    <a href="/admin-logout" class="btn-logout"><i class="fas fa-sign-out-alt"></i> Keluar</a>
                </div>
            </header>
            
            <div class="metrics-grid">
                <div class="metric-card">
                    <div class="metric-title"><i class="fas fa-microchip"></i> Penggunaan CPU</div>
                    <div class="metric-value">{cpu_usage}%</div>
                    <div class="progress-bar-container">
                        <div class="progress-bar" style="width: {cpu_usage}%"></div>
                    </div>
                </div>
                <div class="metric-card">
                    <div class="metric-title"><i class="fas fa-memory"></i> Penggunaan RAM System</div>
                    <div class="metric-value">{mem_percent}%</div>
                    <div style="font-size: 0.8rem; color: var(--text-secondary); margin-bottom: 8px;">
                        Terpakai: {mem_used_mb} MB / Total: {mem_total_mb} MB
                    </div>
                    <div class="progress-bar-container">
                        <div class="progress-bar success" style="width: {mem_percent}%"></div>
                    </div>
                </div>
            </div>
            
            <div class="system-details-card">
                <div class="metric-title" style="margin-bottom: 20px;"><i class="fas fa-circle-info"></i> Rincian Server</div>
                <div class="detail-row">
                    <span class="detail-label">Uptime Server</span>
                    <span class="detail-value" style="color: var(--success-color);">{uptime_str}</span>
                </div>
                <div class="detail-row">
                    <span class="detail-label">RAM Proses (FastAPI)</span>
                    <span class="detail-value">{process_mem_mb} MB</span>
                </div>
                <div class="detail-row">
                    <span class="detail-label">Sistem Operasi</span>
                    <span class="detail-value">{platform.system()} ({platform.machine()})</span>
                </div>
                <div class="detail-row">
                    <span class="detail-label">Versi Python</span>
                    <span class="detail-value">{platform.python_version()}</span>
                </div>
                <div class="detail-row">
                    <span class="detail-label">Folder Sementara (/tmp)</span>
                    <span class="detail-value">{UPLOAD_DIR}</span>
                </div>
            </div>
        </div>
    </body>
    </html>
    """

@app.get("/", response_class=HTMLResponse)
async def get_root(request: Request):
    if not os.environ.get("RENDER"):
        with open("public/index.html", "r", encoding="utf-8") as f:
            return f.read()
            
    token = request.cookies.get("admin_token")
    if token == ADMIN_PASSWORD:
        return get_admin_dashboard_html()
    else:
        return get_login_page_html()

@app.post("/admin-login")
async def admin_login(password: str = Form(...)):
    if password == ADMIN_PASSWORD:
        response = RedirectResponse(url="/", status_code=303)
        response.set_cookie(
            key="admin_token", 
            value=ADMIN_PASSWORD, 
            max_age=86400 * 7, 
            httponly=True,
            samesite="lax"
        )
        return response
    else:
        return get_login_page_html(error="Password salah!")

@app.get("/admin-logout")
async def admin_logout():
    response = RedirectResponse(url="/", status_code=303)
    response.delete_cookie(key="admin_token")
    return response

# Static Files mount at the very end (only when running locally, not on Render)
if not os.environ.get("RENDER"):
    app.mount("/output", StaticFiles(directory="output"), name="output")
    app.mount("/", StaticFiles(directory="public", html=True), name="public")
