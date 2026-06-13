import re
from openpyxl import load_workbook

NAMA_KOLOM_KODE = ['CODE', 'KODE', 'KODE ITEM']
NAMA_KOLOM_NAMA = ['NAME', 'NAMA', 'ITEM NAME', 'NAMA PEMERIKSAAN']
NAMA_KOLOM_KELAS = ['CLASS', 'KELAS']
NAMA_KOLOM_HARGA = ['PRICE UPLOAD', 'PRICE', 'HARGA', 'TARIF', 'BIAYA']

PEMETAAN_KELAS = {
    'OPD': 'OPD', 'ED': 'ED', 'KELAS 3': 'KELAS 3', 'KELAS III': 'KELAS 3',
    'KELAS 2': 'KELAS 2', 'KELAS II': 'KELAS 2', 'KELAS 1': 'KELAS 1',
    'KELAS I': 'KELAS 1', 'VIP': 'VIP', 'VVIP': 'VVIP'
}

def find_column(headers, possible_names):
    for name in possible_names:
        if name in headers:
            return headers[name]
    return None

def parse_price(val):
    if val is None or val == '':
        return None
    if isinstance(val, (int, float)):
        return float(val)
    str_val = str(val).strip()
    if not str_val:
        return None
    clean_str = re.sub(r'[^\d,-]', '', str_val).replace(',', '.')
    try:
        return float(clean_str)
    except ValueError:
        return None

def validate_sheet(input_path):
    errors = []
    warnings = []
    
    wb = load_workbook(input_path, read_only=True)
    if len(wb.worksheets) == 0:
        errors.append("File Excel tidak berisi worksheet.")
        wb.close()
        return {'isValid': False, 'errors': errors, 'warnings': warnings}
        
    sheet = wb.worksheets[0]
    
    # 1. Inspect headers in row 1
    headers = {}
    header_row = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    for col_idx, val in enumerate(header_row, 1):
        if val is not None:
            v_upper = str(val).strip().upper()
            if v_upper:
                headers[v_upper] = col_idx
                
    col_code = find_column(headers, NAMA_KOLOM_KODE)
    col_name = find_column(headers, NAMA_KOLOM_NAMA)
    col_class = find_column(headers, NAMA_KOLOM_KELAS)
    col_price = find_column(headers, NAMA_KOLOM_HARGA)
    
    if not col_code:
        errors.append(f"Kolom Kode tidak ditemukan. Harusnya bernama salah satu dari: {', '.join(NAMA_KOLOM_KODE)}")
    if not col_name:
        errors.append(f"Kolom Nama Pemeriksaan tidak ditemukan. Harusnya bernama salah satu dari: {', '.join(NAMA_KOLOM_NAMA)}")
    if not col_class:
        errors.append(f"Kolom Kelas tidak ditemukan. Harusnya bernama salah satu dari: {', '.join(NAMA_KOLOM_KELAS)}")
    if not col_price:
        errors.append(f"Kolom Harga tidak ditemukan. Harusnya bernama salah satu dari: {', '.join(NAMA_KOLOM_HARGA)}")
        
    if len(errors) > 0:
        wb.close()
        return {'isValid': False, 'errors': errors, 'warnings': warnings}
        
    # 2. Inspect data (first 50 rows)
    unique_classes = set()
    row_count = 0
    
    for row_num, row_cells in enumerate(sheet.iter_rows(min_row=2), 2):
        if row_num > 51:
            break
        row_count += 1
        
        # openpyxl uses 0-based indices in lists, so we subtract 1
        code_val = row_cells[col_code - 1].value
        name_val = row_cells[col_name - 1].value
        price_val = row_cells[col_price - 1].value
        class_val = row_cells[col_class - 1].value
        
        code_str = str(code_val).strip() if code_val is not None else ''
        name_str = str(name_val).strip() if name_val is not None else ''
        class_str = str(class_val).strip().upper() if class_val is not None else ''
        
        if not code_str or not name_str:
            warnings.append(f"Baris {row_num}: Ditemukan baris dengan Kode atau Nama kosong. Baris ini akan dilewati.")
            
        if price_val is not None:
            price_parsed = parse_price(price_val)
            if price_parsed is None and str(price_val).strip():
                errors.append(f"Baris {row_num}: Kolom Harga berisi teks ('{price_val}') yang tidak bisa diubah menjadi angka.")
                
        if class_str:
            unique_classes.add(class_str)
            
    wb.close()
    
    # 3. Check class names consistency
    known_classes = list(PEMETAAN_KELAS.keys())
    for cls in unique_classes:
        if cls not in known_classes:
            warnings.append(f"Ditemukan nama kelas '{cls}' yang tidak dikenal. Harga untuk kelas ini tidak akan diproses secara default.")
            
    return {
        'isValid': len(errors) == 0,
        'errors': errors,
        'warnings': warnings
    }
