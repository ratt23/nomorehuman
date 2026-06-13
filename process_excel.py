import pandas as pd
import numpy as np
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime

def parse_price(val):
    if val is None or pd.isna(val) or val == '':
        return None
    if isinstance(val, (int, float)):
        return float(val)
    str_val = str(val).strip()
    if not str_val:
        return None
    # cleanStr = strVal.replace(/[^\d,-]/g, '').replace(',', '.');
    clean_str = re.sub(r'[^\d,-]', '', str_val).replace(',', '.')
    try:
        return float(clean_str)
    except ValueError:
        return None

def process_and_diagnose_sheet(input_path, mappings, class_map, selected_sheet, filter_config):
    # Read Excel using pandas
    df = pd.read_excel(input_path, sheet_name=selected_sheet, dtype=str)
    
    # We want to match headers
    col_code = mappings.get('kode')
    col_name = mappings.get('nama')
    col_class = mappings.get('kelas')
    col_price = mappings.get('harga')
    
    if not all(col in df.columns for col in [col_code, col_name, col_class, col_price]):
        raise Exception("Pemetaan kolom tidak valid.")
        
    # Check if we have a filter
    use_filter = filter_config and filter_config.get('column') and len(filter_config.get('values', [])) > 0
    col_filter = filter_config.get('column') if use_filter else None
    filter_values = [str(v).upper() for v in filter_config.get('values', [])] if use_filter else []
    
    rejected_rows_sample = []
    REJECTED_SAMPLE_SIZE = 10
    total_rows_read = 0
    total_rows_filtered = 0
    
    grouped = {}
    
    for idx, row in df.iterrows():
        total_rows_read += 1
        
        # Apply filter
        if use_filter and col_filter in df.columns:
            cell_value = str(row[col_filter]).strip() if not pd.isna(row[col_filter]) else ''
            if cell_value.upper() not in filter_values:
                if len(rejected_rows_sample) < REJECTED_SAMPLE_SIZE:
                    rejected_rows_sample.append({
                        'kode': str(row[col_code]).strip() if not pd.isna(row[col_code]) else '',
                        'nama': str(row[col_name]).strip() if not pd.isna(row[col_name]) else '',
                        'alasan': f"Nilai '{cell_value}' tidak cocok filter"
                    })
                total_rows_filtered += 1
                continue
                
        code = str(row[col_code]).strip() if not pd.isna(row[col_code]) else ''
        name = str(row[col_name]).strip() if not pd.isna(row[col_name]) else ''
        
        if not code or not name:
            continue
            
        kelas_from_file = str(row[col_class]).strip().upper() if not pd.isna(row[col_class]) else ''
        price_raw = row[col_price]
        price = parse_price(price_raw)
        
        key = f"{code}|{name}"
        if key not in grouped:
            grouped[key] = {
                'code': code,
                'name': name,
                'OPD': None,
                'ED': None,
                'KELAS 3': None,
                'KELAS 2': None,
                'KELAS 1': None,
                'VIP': None,
                'VVIP': None
            }
            
        target_col = class_map.get(kelas_from_file)
        if target_col and target_col != 'ignore':
            grouped[key][target_col] = price
            
    all_accepted_rows = list(grouped.values())
    accepted_rows_sample = all_accepted_rows[:10]
    
    summary = {
        'totalRowsRead': total_rows_read,
        'totalRowsFiltered': total_rows_filtered,
        'totalRowsProcessed': total_rows_read - total_rows_filtered,
        'uniqueItemCount': len(all_accepted_rows)
    }
    
    return {
        'summary': summary,
        'rejectedRowsSample': rejected_rows_sample,
        'acceptedRowsSample': accepted_rows_sample,
        'allAcceptedRows': all_accepted_rows
    }

def create_final_excel(rows, output_dir):
    wb = Workbook()
    ws = wb.active
    ws.title = "Buku Tarif LAB"
    
    headers = [
        'Kode', 'Nama Pemeriksaan', 'OPD', 'ED', 
        'KELAS 3', 'KELAS 2', 'KELAS 1', 'VIP', 'VVIP'
    ]
    
    ws.append(headers)
    
    # Column styling
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        
    # Append rows
    for r in rows:
        row_data = [
            r.get('code', ''),
            r.get('name', ''),
            r.get('OPD'),
            r.get('ED'),
            r.get('KELAS 3'),
            r.get('KELAS 2'),
            r.get('KELAS 1'),
            r.get('VIP'),
            r.get('VVIP')
        ]
        ws.append(row_data)
        
    # Format number columns (columns 3 to 9) to standard number format #,##0
    for row_idx in range(2, len(rows) + 2):
        for col_idx in range(3, 10):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell.number_format = '#,##0'
                
    # Auto-adjust column widths
    column_widths = {
        1: 15, 2: 45, 3: 15, 4: 15, 5: 15, 6: 15, 7: 15, 8: 15, 9: 15
    }
    for col_idx, width in column_widths.items():
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
        
    timestamp = datetime.now().strftime('%Y-%m-%dT%H-%M-%S')
    out_file_name = f"Buku_Tarif_LAB_{timestamp}.xlsx"
    out_path = os.path.join(output_dir, out_file_name)
    wb.save(out_path)
    return out_path
