from openpyxl import load_workbook
import os
from datetime import datetime

def get_combined_headers(sheet, header_row_count):
    if header_row_count <= 0:
        return []
    
    rows_data = []
    max_cols = 0
    for r_idx in range(1, header_row_count + 1):
        row_cells = [cell.value for cell in sheet[r_idx]]
        max_cols = max(max_cols, len(row_cells))
        rows_data.append(row_cells)
        
    if header_row_count == 1:
        headers = []
        for col_idx, val in enumerate(rows_data[0], 1):
            val_str = str(val).strip() if val is not None else ''
            headers.append(val_str if val_str else f"Kolom_{col_idx}")
        return headers
        
    header_matrix = []
    for r in range(header_row_count):
        row_vals = []
        for c in range(max_cols):
            val = rows_data[r][c] if c < len(rows_data[r]) else None
            row_vals.append(str(val).strip() if val is not None else '')
        header_matrix.append(row_vals)
        
    final_headers = []
    for c in range(max_cols):
        context = ''
        header_parts = []
        for r in range(header_row_count):
            cell_value = header_matrix[r][c]
            if cell_value:
                context = cell_value
            if r == header_row_count - 1 or (r + 1 < header_row_count and header_matrix[r+1][c]):
                if context and context not in header_parts:
                    header_parts.append(context)
        final_header = ' - '.join(header_parts)
        final_headers.append(final_header if final_header else f"Kolom_{c + 1}")
        
    return final_headers

def inspect_source_file(file_path, config=None):
    if config is None:
        config = {'headerRowCount': 'auto'}
        
    wb = load_workbook(file_path, read_only=True)
    result = {
        'sheets': [],
        'headersBySheet': {},
        'detectedHeaderRows': {}
    }
    
    if len(wb.worksheets) == 0:
        wb.close()
        raise Exception("File Excel tidak berisi sheet yang valid atau kosong.")
        
    for sheet in wb.worksheets:
        sheet_name = sheet.title
        result['sheets'].append(sheet_name)
        
        header_row_count = 1
        normal_wb = load_workbook(file_path)
        normal_sheet = normal_wb[sheet_name]
        
        if config.get('headerRowCount') == 'auto':
            row1_has_merge = False
            row2_has_merge = False
            
            for merged_range in normal_sheet.merged_cells.ranges:
                min_col, min_row, max_col, max_row = merged_range.bounds
                if min_row <= 1 <= max_row:
                    row1_has_merge = True
                if min_row <= 2 <= max_row:
                    row2_has_merge = True
                    
            if row1_has_merge and row2_has_merge:
                header_row_count = 3
            elif row1_has_merge:
                header_row_count = 2
        else:
            try:
                header_row_count = int(config.get('headerRowCount', 1))
            except ValueError:
                header_row_count = 1
                
        result['detectedHeaderRows'][sheet_name] = header_row_count
        result['headersBySheet'][sheet_name] = get_combined_headers(normal_sheet, header_row_count)
        normal_wb.close()
        
    wb.close()
    return result

def inspect_template_file(file_path):
    wb = load_workbook(file_path, read_only=True)
    if len(wb.worksheets) == 0:
        wb.close()
        raise Exception("File template tidak valid atau tidak berisi sheet.")
        
    sheet = wb.worksheets[0]
    headers = []
    
    row1 = next(sheet.iter_rows(max_row=1))
    for cell in row1:
        if cell.value is not None:
            val_str = str(cell.value).strip()
            if val_str:
                headers.append(val_str)
                
    wb.close()
    if len(headers) == 0:
        raise Exception("Template tidak berisi header di baris pertama.")
    return headers

def build_report(file_path, config):
    wb = load_workbook(file_path, data_only=True)
    source_sheet_name = config.get('selectedSheet')
    if source_sheet_name not in wb.sheetnames:
        wb.close()
        raise Exception(f'Sheet "{source_sheet_name}" tidak ditemukan.')
        
    source_sheet = wb[source_sheet_name]
    header_row_count = int(config.get('headerRowCount', 1))
    source_headers = get_combined_headers(source_sheet, header_row_count)
    
    from openpyxl import Workbook
    new_wb = Workbook()
    out_sheet = new_wb.active
    out_sheet.title = "Rekap Laporan"
    
    mappings = config.get('mappings', [])
    
    output_headers = [m.get('outputHeader') for m in mappings]
    out_sheet.append(output_headers)
    
    from openpyxl.styles import Font
    header_font = Font(bold=True)
    for col_idx in range(1, len(output_headers) + 1):
        out_sheet.cell(row=1, column=col_idx).font = header_font
        
    unique_key_set = set()
    diagnostics = {
        'totalRowsRead': 0,
        'rowsAdded': 0,
        'rowsSkipped_DuplicateKey': 0
    }
    
    unique_key_col = config.get('uniqueKeyColumn')
    
    for row_num in range(header_row_count + 1, source_sheet.max_row + 1):
        diagnostics['totalRowsRead'] += 1
        
        if unique_key_col:
            try:
                key_cell_idx = source_headers.index(unique_key_col) + 1
                key_value = str(source_sheet.cell(row=row_num, column=key_cell_idx).value or '').strip()
                if key_value:
                    if key_value in unique_key_set:
                        diagnostics['rowsSkipped_DuplicateKey'] += 1
                        continue
                    unique_key_set.add(key_value)
            except ValueError:
                pass
                
        new_row_data = []
        row_has_data = False
        
        for mapping in mappings:
            source_header = mapping.get('sourceHeader')
            val = None
            try:
                source_cell_idx = source_headers.index(source_header) + 1
                cell = source_sheet.cell(row=row_num, column=source_cell_idx)
                val = cell.value
                if val is not None and str(val).strip() != '':
                    row_has_data = True
            except ValueError:
                pass
            new_row_data.append(val)
            
        if row_has_data:
            out_sheet.append(new_row_data)
            
    diagnostics['rowsAdded'] = out_sheet.max_row - 1
    
    output_dir = os.path.join(os.path.dirname(__file__), 'output')
    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime('%Y-%m-%dT%H-%M-%S')
    out_file_name = f"Laporan_Kustom_{timestamp}.xlsx"
    out_path = os.path.join(output_dir, out_file_name)
    
    new_wb.save(out_path)
    new_wb.close()
    wb.close()
    
    return out_path, diagnostics
